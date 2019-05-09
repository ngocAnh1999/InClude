#!/usr/bin/env python
# -*- coding: utf-8 -*-

from flask_sqlalchemy import SQLAlchemy
from models import Mon_an, Thanh_phan, Van_hoa, Cach_cb, _mua, Meovaobep, Meovat
import openpyxl, requests, pymysql
from api import sqldb


path = "/home/ngocanh/Downloads/Database_Congthucnauan.xlsx"
wb = openpyxl.load_workbook(path)

def checkNull(ten_mon, image, cong_thuc, nguyen_lieu):
        if(ten_mon is None or image is None or cong_thuc is None or nguyen_lieu is None):
                print("Some data is null, please check again at ten_mon = {}".format(ten_mon))
                return False;
        return True

def importMua():
        sheet = wb.get_sheet_by_name("Bang_mua_dip_le")
        max_col = sheet.max_column
        max_row = sheet.max_row
        for i in range(0, max_row - 1):
                cell = sheet.cell(row = i+2, column = 2)
                mua = _mua()
                if(mua.query.filter(mua.name == cell.value).first() is None):
                        mua.addMua(cell.value)
                        print(cell.value + "++++++++++++successfully")
                
def importThanh_phan():
        sheet = wb.get_sheet_by_name("Bang_thanh_phan")
        max_col = sheet.max_column
        max_row = sheet.max_row
        for i in range(0, max_row - 1):
                cell = sheet.cell(row = i+2, column = 2)
                thanhphan = Thanh_phan()
                if(thanhphan.query.filter(thanhphan.name == cell.value).first() is None):
                        thanhphan.addThanhphan(cell.value)
                        print(cell.value + "++++++++++++successfully")

def importVan_hoa():
        sheet = wb.get_sheet_by_name("Bang_van_hoa")
        max_col = sheet.max_column
        max_row = sheet.max_row
        for i in range(0, max_row - 1):
                cell = sheet.cell(row = i+2, column = 2)
                vanhoa = Van_hoa()
                if(vanhoa.query.filter(vanhoa.name == cell.value).first() is None):
                        vanhoa.addVanhoa(cell.value)
                        print(cell.value + "++++++++++++successfully")

def importCachchebien():
        sheet = wb.get_sheet_by_name("Bang_cach_che_bien")
        max_col = sheet.max_column
        max_row = sheet.max_row
        for i in range(0, max_row - 1):
                cell = sheet.cell(row = i+2, column = 2)
                cachchebien = Cach_cb()
                if(cachchebien.query.filter(cachchebien.name == cell.value).first() is None):
                        cachchebien.addCachchebien(cell.value)
                        print(cell.value + "++++++++++++successfully")

def importMon_an():
        sheet = wb.get_sheet_by_name("Bang_mon_an")
        max_col = sheet.max_column
        max_row = sheet.max_row
        print max_col
        print max_row
        for i in range(0, max_row - 1):
                ten_mon = sheet.cell(row = i+2, column = 2).value
                image = sheet.cell(row = i+2, column = 3).value
                cong_thuc = sheet.cell(row = i+2, column = 4).value
                nguyen_lieu = sheet.cell(row = i+2, column = 5).value
                ma_nl = sheet.cell(row = i+2, column = 6).value
                ma_cach_cb = sheet.cell(row = i+2, column = 7).value
                ma_vh = sheet.cell(row = i+2, column = 8).value
                ma_mua = sheet.cell(row = i+2, column = 9).value
                video = sheet.cell(row = i+2, column = 10).value
                
                monan = Mon_an()
                results = monan.query.with_entities(Mon_an.ten_mon).filter(monan.ten_mon==ten_mon).first()
                print results
                # if(len(results) == 0 and checkNull(ten_mon, image, cong_thuc, nguyen_lieu)):
                #         print(ten_mon + "++++++++++++successfully")
                #         monan.addMonan(ten_mon, image, cong_thuc, nguyen_lieu, ma_nl, ma_vh, ma_cach_cb, ma_mua, video)

def importMeovaobep():
        sheet = wb.get_sheet_by_name("Bang_meo_vao_bep")
        max_col = sheet.max_column
        max_row = sheet.max_row
        for i in range(0, max_row - 1):
                name = sheet.cell(row = i+2, column = 3).value
                mo_ta = sheet.cell(row = i+2, column = 4).value
                image = sheet.cell(row = i+2, column = 5).value
                id_meo = sheet.cell(row = i+2, column = 2).value
                meovaobep = Meovaobep()
                if(meovaobep.query.filter(meovaobep.name == name).all() is None):
                        print(name + "++++++++++ successfully")
                        meovaobep.addMeovaobep(name, mo_ta, image, id_meo)
                
def importMeovat():
        sheet = wb.get_sheet_by_name("Bang_meo_vat")
        max_col = sheet.max_column
        max_row = sheet.max_row
        print(max_col)
        for i in range(0, max_row - 1):
                name = sheet.cell(row = i+2, column = 2).value
                mo_ta = sheet.cell(row = i+2, column = 3).value
                print(name)
                meovat = Meovat()
                if(meovat.query.filter(meovat.name == name).all() is None):
                        print(name + "++++++++++ successfully")
                        meovat.addMeovat(name, mo_ta)

if __name__ == '__main__':
        # print("import data for table Mua:")
        # importMua()
        # print("import data for table Thanh_phan:")
        # importThanh_phan()
        # print("import data for table Van_hoa:")
        # importVan_hoa()
        # print("import data for table Cach_che_bien:")
        # importCachchebien()
        print("import data for table Mon_an:")
        importMon_an()
        # print("import data for table Meo_vat:")
        # importMeovat()
        # print("import data for table Meo_vao_bep:")
        # importMeovaobep() 
        