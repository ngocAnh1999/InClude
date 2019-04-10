#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl, requests

path = "/home/ngocanh/Downloads/bang_ble.xlsx"

wb = openpyxl.load_workbook(path)
sheet = wb.get_sheet_by_name("phuongtiengiaothongvi")
max_row = sheet.max_row
url = "http://api.openfpt.vn/text2speech/v5"
headers = {'voice': 'banmai','api_key': '274ce4dd5814412e8fcd4ca0fa752bdf'}
for i in range(2, max_row + 1):
    cell = sheet.cell(row = i, column = 2)
    data = cell.value
    r = requests.post(url, headers=headers, data=data)
    print(r.json())