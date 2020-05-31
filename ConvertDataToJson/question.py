import json
from openpyxl import load_workbook
wb = load_workbook('./ONTHI_GPLX.xlsx')

data = []

if __name__ == '__main__':
    questions = wb["Question"]
    answers = wb["Answer"]
    typeQ = wb["TypeQuestion"]
    row_q = questions.max_row
    row_a = answers.max_row
    row_tq = typeQ.max_row
    count_q = 1
    count_a = 1
    for ti in range(1, row_tq):
        _id = typeQ["A"][ti].value
        _type = typeQ["B"][ti].value
        _desciption = typeQ["C"][ti].value
        _questions = []
        for qi in range(count_q, row_q):
            _answers = []
            id_type = questions["F"][count_q].value
            if(id_type == _id):
                id_q = questions["A"][qi].value
                content_q = questions["B"][qi].value
                img = questions["C"][qi].value
                explain = questions["D"][qi].value
                top50 = questions["E"][qi].value
                for ai in range(count_a, row_a):
                    id_qa = answers["C"][ai].value
                    if(id_qa == id_q):
                        id_a = answers["A"][ai].value
                        content_a = answers["B"][ai].value
                        check = answers["D"][ai].value
                        _answers.append({
                            "id": id_a,
                            "content": content_a,
                            "check": check
                        })
                    else:
                        count_a = ai
                        break
                _questions.append({
                    "id": id_q,
                    "content": content_q,
                    "img": img,
                    "answers": _answers,
                    "explain": explain,
                    "top50": top50
                })
            else:
                count_q = qi
                break


        data.append({
            "id": _id,
            "type": _type,
            "description": _desciption,
            "question": _questions
        })
    
            
    with open('question.json', 'w') as outfile:
        json.dump(data, outfile, ensure_ascii=False)