# -*- coding: utf-8 -*-
import json
from openpyxl import load_workbook
wb = load_workbook('./TraficSignDatas.xlsx')

data = []

if __name__ == '__main__':
    sheet1 = wb['Sheet1']
    max_row = sheet1.max_row
    count = 0
    # key = 0
    for i in range(2, max_row+1):
        pre_type = sheet1['B'][i-2].value
        _type = sheet1['B'][i-1].value
        next_type = None
        if(i < max_row):
            next_type = sheet1['B'][i].value
        _id = sheet1['A'][i-1].value
        _title = sheet1['C'][i-1].value
        img_link = sheet1['D'][i-1].value
        _content = sheet1['E'][i-1].value
        if(i == 2 or _type != pre_type):
            print(f"count = {count}")
            data.append({
                "type": _type,
                'body' : [
                    {
                        'id': _id,
                        'title': _title,
                        'img_link': img_link,
                        'content': _content
                    }
                ]
            })
            # key = 1
        else :
            # print(f"key = {key}")
            data[count]['body'].append({
                'id': _id,
                'title': _title,
                'img_link': img_link,
                'content': _content
            })
            # key = key + 1
        if(_type != next_type):
            count = count + 1
    with open('traficSigns.json', 'w') as outfile:
        json.dump(data, outfile, ensure_ascii=False)
    
    